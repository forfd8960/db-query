"""Excel export implementation for query results.

This module provides Excel XLSX export functionality with native type formatting,
proper date handling, and cell formatting.
"""

import io
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from src.utils.filename import disambiguate_column_names

from .base import BaseExporter


class ExcelExporter(BaseExporter):
    """Export query results to Excel XLSX format with native types."""

    def export(self, columns: list[str], rows: list[dict[str, Any]]) -> bytes:
        """Export data to Excel XLSX format.

        Args:
            columns: List of column names (order preserved)
            rows: List of row dictionaries with column names as keys

        Returns:
            Excel XLSX data as bytes

        Notes:
            - Creates worksheet named "Query Results"
            - Headers in row 1 with bold formatting
            - Data rows start from row 2
            - Numbers exported as native Excel numbers
            - Dates exported with Excel date formatting
            - Booleans exported as Excel TRUE/FALSE
            - Null values exported as empty cells (None)
            - Handles duplicate column names with _1, _2 suffixes
        """
        # Disambiguate duplicate column names
        unique_columns = disambiguate_column_names(columns)
        
        # Create workbook and worksheet
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Query Results"

        # Write column headers in row 1 with bold formatting
        header_font = Font(bold=True)
        for col_idx, column in enumerate(unique_columns, start=1):
            cell = worksheet.cell(row=1, column=col_idx, value=column)
            cell.font = header_font

        # Write data rows starting from row 2
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, column in enumerate(columns, start=1):
                value = row.get(column)
                cell = worksheet.cell(row=row_idx, column=col_idx)

                # Handle different value types
                if value is None:
                    # Null values become empty cells
                    cell.value = None
                elif isinstance(value, bool):
                    # Boolean values as Excel TRUE/FALSE
                    cell.value = value
                elif isinstance(value, (int, float)):
                    # Numbers as native Excel numbers
                    cell.value = value
                elif isinstance(value, datetime):
                    # Datetime with Excel formatting
                    cell.value = value
                    cell.number_format = "yyyy-mm-dd hh:mm:ss"
                elif isinstance(value, date):
                    # Date with Excel formatting
                    cell.value = value
                    cell.number_format = "yyyy-mm-dd"
                else:
                    # Everything else as string
                    cell.value = str(value)

        # Save workbook to BytesIO buffer
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        return buffer.read()
